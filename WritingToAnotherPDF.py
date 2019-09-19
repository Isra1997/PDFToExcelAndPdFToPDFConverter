import openpyxl
from reportlab.pdfgen import canvas
from PyPDF2 import PdfFileMerger
import os
from PyPDF2 import PdfFileWriter, PdfFileReader
from reportlab.lib.pagesizes import portrait
from io import StringIO

class Invoice:
    def __init__(self,CustomerNumber,POReferance,OurReference,InvoiceNummber,VATax):
        self.CustomerNumber = CustomerNumber
        self.POReferance = POReferance
        self.OurReference=OurReference
        self.InvoiceNummber=InvoiceNummber
        self.VATax=VATax

# function that reads cells from an excel sheet
def reading_From_excel(Sheet_Path):
    wb= openpyxl.load_workbook(Sheet_Path)
    sheet=wb.active
    ynum=sheet.max_row
    for y  in range(1,ynum+1):
        CustomerNumber=sheet.cell(row=y, column=1)
        POReferance=sheet.cell(row=y, column=2)
        OurReference=sheet.cell(row=y, column=3)
        InvoiceNummber=sheet.cell(row=y, column=4)
        VATax=sheet.cell(row=y, column=5)
        invo=WriteToPDF(CustomerNumber.value, POReferance.value, OurReference.value, InvoiceNummber.value, VATax.value,y)

# function that writes the read output from the excel sheet and writes it to a pdf
def WriteToPDF(CustomerNumber, POReferance, OurReference, InvoiceNummber, VAT,num):
    c= canvas.Canvas("Inv"+str(num)+".pdf")
    c.drawString(70,750, "Invoice Nummber:   "+str(InvoiceNummber))
    c.drawString(70,720, "Customer Number:   "+str(CustomerNumber))
    c.drawString(70,690, "Our Reference:   "+str(OurReference))
    c.drawString(70,600, "PO Referance:   "+str(POReferance))
    c.drawString(70,100, "VAT:   "+str(VAT))
    c.save()


def main():
    reading_From_excel("/Users/israragheb/Desktop/demo.xlsx")

if __name__=='__main__':
    main()
