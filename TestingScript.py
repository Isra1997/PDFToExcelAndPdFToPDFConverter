import pytesseract
from pdf2image import convert_from_path
import tempfile
import os
from pytesseract import image_to_string
from PIL import Image
import PyPDF2
import openpyxl

class Invoice:
    def __init__(self,CustomerNumber,POReferance,OurReference,InvoiceNummber,VATax):
        self.CustomerNumber = CustomerNumber
        self.POReferance = POReferance
        self.OurReference=OurReference
        self.InvoiceNummber=InvoiceNummber
        self.VATax=VATax

# function PDF -> image -> text
def  StartingFunction(PDF_file):
# PDF_file = "/Users/israragheb/Desktop/FileOfInvocies/2.pdf"
    images_from_path=convert_from_path(PDF_file, dpi=500)
    x=0
    for image in images_from_path:
        num=str(x)
        images_from_path[x].save(num+'.jpg', 'JPEG')
        file=open('PDFtest.txt','a')
        image=Image.open(num+'.jpg')
        file.write(image_to_string(image))
        x=x+1
    file.close()

def SplitingFunction(txt_filepath):
    file =open(txt_filepath)
    content=file.read()
    contentFinal=content.splitlines()
    InvoiceNummberArray=contentFinal[106].split()
    InvoiceNummber=InvoiceNummberArray[4]
    print(InvoiceNummber)
    CustomerNumberArray=contentFinal[108].split()
    CustomerNumber=CustomerNumberArray[4]
    print(CustomerNumber)
    OurReferenceArray=contentFinal[109].split()
    OurReference=OurReferenceArray[4]
    print(OurReference)
    POReferanceArray=contentFinal[141].split()
    POReferance=POReferanceArray[0]
    print(POReferance)
    VATArray=contentFinal[248].split()
    VAT=VATArray[1]
    print(VAT)
    invoice=Invoice(CustomerNumber,POReferance,OurReference,InvoiceNummber,VAT)
    return invoice

def WriteToExcel(invoice,excelFilePath):
    wb = openpyxl.Workbook()
    sheet = wb.active
    c0 = sheet.cell(row = 1, column = 1)
    c0.value = "CustomerNumber"
    c1= sheet.cell(row = 1, column = 2)
    c1.value="POReferance"
    c2=sheet.cell(row=1, column=3)
    c2.value="OurReference"
    c3=sheet.cell(row=1, column=4)
    c3.value="InvoiceNummber"
    c4=sheet.cell(row=1, column=5)
    c4.value="VAT"
    c4=sheet.cell(row=4, column=1)
    c4.value=invoice.CustomerNumber
    c5=sheet.cell(row=4, column=2)
    c5.value=invoice.POReferance
    c6=sheet.cell(row=4, column=3)
    c6.value=invoice.OurReference
    c7=sheet.cell(row=4, column=4)
    c7.value=invoice.InvoiceNummber
    c8=sheet.cell(row=4, column=5)
    c8.value=invoice.VATax
    wb.save(excelFilePath)


def main():
    # StartingFunction("/Users/israragheb/Desktop/FileOfInvocies/2.pdf")
    WriteToExcel(SplitingFunction("PDFtest.txt"),"/Users/israragheb/Desktop/demo.xlsx")

if __name__=='__main__':
    main()
