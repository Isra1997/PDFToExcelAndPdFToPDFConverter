# reading the pdf file
from tika import parser
import openpyxl
import os
import pytesseract
from pdf2image import convert_from_path
import tempfile
from pytesseract import image_to_string
from PIL import Image
import PyPDF2
import sys

class Invoice:
    def __init__(self,CustomerNumber,POReferance,OurReference,InvoiceNummber,VATax):
        self.CustomerNumber = CustomerNumber
        self.POReferance = POReferance
        self.OurReference=OurReference
        self.InvoiceNummber=InvoiceNummber
        self.VATax=VATax

#function for looping on all Invoices
def IntiateArrayOfInvoice(filepath):
    arrary=[]
    directory = os.fsencode(filepath)
    log=open('parsing.txt','a')
    for file in os.listdir(directory):
        try:
            filename = os.fsdecode(file)
            if filename.endswith(".pdf"):
                raw = parser.from_file(filepath+'/'+filename)
                # print(raw['content'])
                line=raw['content'].splitlines()
                InvoiceNummberArray=line[43].split()
                CustomerNumberArray=line[45].split()
                OurReferenceArray=line[47].split()
                InvoiceNummber=InvoiceNummberArray[2]
                # print(InvoiceNummber)
                CustomerNumber=CustomerNumberArray[2]
                # print(CustomerNumber)
                POReferance=line[59]
                # print(POReferance)
                OurReference=OurReferenceArray[2]
                # print(OurReference)
                # print('/')
                #vat
                inv=Invoice(CustomerNumber,POReferance,OurReference,InvoiceNummber,"")
                arrary.append(inv)
            else:
                continue
        except AttributeError:
            # log.write(filepath+'/'+filename)
            invo=SplitingFunction(StartingFunction(filepath+'/'+filename))
            arrary.append(invo)
    log.close()
    return arrary
    # return an array of file path also

# function for writing to a excel sheet
def WriteToExcel(arrary,excelFilePath):
    wb = openpyxl.Workbook()
    sheet = wb.active
    c0 = sheet.cell(row = 1, column = 1)
    c0.value = "CustomerNumber"
    c1= sheet.cell(row = 1, column = 2)
    c1.value="POReferance"
    c2=sheet.cell(row=1, column=3)
    c2.value="OurReference"
    c3=sheet.cell(row=1, column=4)
    c3.value="InvoiceNummber"
    c4=sheet.cell(row=1, column=5)
    c4.value="VAT"
    r=2
    for invoice in arrary:
        c4=sheet.cell(row=r, column=1)
        c4.value=invoice.CustomerNumber
        c5=sheet.cell(row=r, column=2)
        c5.value=invoice.POReferance
        c6=sheet.cell(row=r, column=3)
        c6.value=invoice.OurReference
        c7=sheet.cell(row=r, column=4)
        c7.value=invoice.InvoiceNummber
        c8=sheet.cell(row=r, column=5)
        c8.value=invoice.VATax
        r=r+1
    wb.save(excelFilePath)

# function that is used convert pdf-> images -> text
def  StartingFunction(PDF_file):
# PDF_file = "/Users/israragheb/Desktop/FileOfInvocies/2.pdf"
    images_from_path=convert_from_path(PDF_file, dpi=500)
    x=0
    for image in images_from_path:
        num=str(x)
        images_from_path[x].save(num+'.jpg', 'JPEG')
        file=open('PDFtest.txt','a')
        image=Image.open(num+'.jpg')
        file.write(image_to_string(image))
        x=x+1
    file.close()
    return 'PDFtest.txt'

# function that splits the text from image pdf
def SplitingFunction(txt_filepath):
    file =open(txt_filepath)
    content=file.read()
    contentFinal=content.splitlines()
    InvoiceNummberArray=contentFinal[106].split()
    InvoiceNummber=InvoiceNummberArray[4]
    # print(InvoiceNummber)
    CustomerNumberArray=contentFinal[108].split()
    CustomerNumber=CustomerNumberArray[4]
    # print(CustomerNumber)
    OurReferenceArray=contentFinal[109].split()
    OurReference=OurReferenceArray[4]
    # print(OurReference)
    POReferanceArray=contentFinal[141].split()
    POReferance=POReferanceArray[0]
    # print(POReferance)
    VATArray=contentFinal[248].split()
    VAT=VATArray[1]
    # print(VAT)
    invoice=Invoice(CustomerNumber,POReferance,OurReference,InvoiceNummber,VAT)
    return invoice



def main():
    # print("Out")
    # one array is send to WriteToExcel
    WriteToExcel(IntiateArrayOfInvoice("/Users/israragheb/FileOfInvocies"),"/Users/israragheb/Desktop/demo.xlsx")


if __name__=='__main__':
    main()
